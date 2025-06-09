# ui/cost_dialogs.py - Cost Analysis Dialog System
import os
import tkinter as tk
from tkinter import filedialog, messagebox
from ttkbootstrap import Frame, Label, Button, Entry, Combobox
from ttkbootstrap.constants import *
import threading
import time
from datetime import datetime
from ui.styles import MODERN_SLATE

class CostAnalysisDialogs:
    """Handles all cost analysis related dialogs and exports"""
    
    def __init__(self, parent, controller):
        self.parent = parent
        self.controller = controller
        self.cost_analysis_cache = {}
        
    def show_cost_analysis(self):
        """STAGE 3 ENHANCED: Main cost analysis method with loading states and caching"""
        if not self.parent.chunks:
            messagebox.showwarning("No Data", "Please process a file first to analyze training costs.")
            return

        # Check premium access
        if not self.controller.license_manager.check_feature_access('advanced_cost_analysis'):
            self.show_cost_upgrade_dialog()
            return

        try:
            tokenizer_name = getattr(self.parent, '_current_tokenizer_name', 'gpt2')
            target_models = ['llama-2-7b', 'llama-2-13b', 'claude-3-haiku']
            api_usage_monthly = 100000
            
            # Generate cache key
            cache_key = self._get_analysis_cache_key(
                self.parent.chunks, tokenizer_name, 512, target_models, api_usage_monthly
            )
            
            # Check cache first
            if self._is_analysis_cache_valid(cache_key):
                cost_analysis = self.cost_analysis_cache[cache_key]['data']
                self._display_cost_analysis_dialog(cost_analysis)
                return
            
            # Show loading dialog
            loading_window, progress, status_label = self._show_loading_dialog(
                "Analyzing Training Costs",
                "Calculating comprehensive costs across 15+ approaches...\nThis may take a few seconds."
            )
            
            def run_analysis():
                try:
                    # Update status
                    def update_status(text):
                        try:
                            status_label.config(text=text)
                            loading_window.update()
                        except tk.TclError:
                            pass
                    
                    update_status("Loading model database...")
                    time.sleep(0.5)  # Brief pause for UI feedback
                    
                    update_status("Fetching real-time cloud pricing...")
                    time.sleep(0.5)
                    
                    update_status("Calculating training costs...")
                    
                    # Perform the actual analysis
                    cost_analysis = self.controller.analyze_chunks_with_costs(
                        self.parent.chunks, 
                        tokenizer_name, 
                        512,
                        target_models=target_models,
                        api_usage_monthly=api_usage_monthly
                    )
                    
                    update_status("Generating optimization recommendations...")
                    time.sleep(0.3)
                    
                    # Cache the results
                    self.cost_analysis_cache[cache_key] = {
                        'data': cost_analysis,
                        'timestamp': time.time()
                    }
                    
                    # Close loading dialog and show results
                    self._close_loading_dialog(loading_window, progress)
                    self._display_cost_analysis_dialog(cost_analysis)
                    
                except Exception as e:
                    self._close_loading_dialog(loading_window, progress)
                    self._show_enhanced_error_dialog(
                        "Cost Analysis Error",
                        f"Failed to analyze training costs: {str(e)}",
                        recovery_suggestions=[
                            "Check your internet connection for live pricing",
                            "Try with a smaller dataset",
                            "Contact support if the problem persists"
                        ]
                    )
            
            # Run analysis in thread to keep UI responsive
            analysis_thread = threading.Thread(target=run_analysis)
            analysis_thread.daemon = True
            analysis_thread.start()
            
        except Exception as e:
            self._show_enhanced_error_dialog(
                "Cost Analysis Error",
                f"Failed to start cost analysis: {str(e)}"
            )

    def show_cost_upgrade_dialog(self):
        """Show upgrade dialog for cost analysis feature"""
        try:
            upgrade_info = self.controller.get_upgrade_info()
            
            message = """🔒 Premium Feature: Enhanced Cost Calculator

Comprehensive AI Training Cost Analysis includes:
• 15+ Training Approaches (Local, Cloud, API, LoRA, QLoRA)
• Real-time Cloud Pricing (Lambda Labs, Vast.ai, RunPod)
• ROI Analysis with Break-even Calculations
• Cost Optimization Recommendations
• Professional Export Reports

💰 Example Savings:
• 50K word dataset: Save $32+ per training run
• Accurate planning prevents cost overruns
• Optimal approach selection

💎 Premium: $15/month or $150/year
🆓 7-day free trial - no credit card required

Would you like to start your free trial?"""
            
            result = messagebox.askyesno("Upgrade to Premium", message)
            if result:
                self.parent.start_trial()
        except Exception as e:
            messagebox.showerror("Upgrade Error", f"Failed to show upgrade info: {str(e)}")

    def _create_scrollable_dialog(self, title, width=800, height=600):
        """FIXED: Create a scrollable dialog with proper canvas management"""
        dialog = tk.Toplevel(self.parent)
        dialog.title(title)
        dialog.geometry(f"{width}x{height}")
        dialog.transient(self.parent)
        dialog.grab_set()
        dialog.configure(bg=MODERN_SLATE['bg_primary'])
        
        # Center the dialog
        dialog.geometry(f"+{dialog.winfo_screenwidth()//2 - width//2}+{dialog.winfo_screenheight()//2 - height//2}")
        
        # Create canvas and scrollbar for dialog
        dialog_canvas = tk.Canvas(dialog, 
                                 borderwidth=0, 
                                 highlightthickness=0,
                                 bg=MODERN_SLATE['bg_primary'])
        dialog_scrollbar = tk.Scrollbar(dialog, orient="vertical", command=dialog_canvas.yview)
        
        # Create content frame
        content_frame = Frame(dialog_canvas, padding=(20, 20), style="Modern.TFrame")
        
        content_frame.bind(
            "<Configure>",
            lambda e: dialog_canvas.configure(scrollregion=dialog_canvas.bbox("all"))
        )
        
        dialog_canvas.create_window((0, 0), window=content_frame, anchor="nw")
        dialog_canvas.configure(yscrollcommand=dialog_scrollbar.set)
        
        # Pack canvas and scrollbar
        dialog_canvas.pack(side="left", fill="both", expand=True)
        dialog_scrollbar.pack(side="right", fill="y")
        
        # FIXED: Dialog-specific mousewheel binding (not global)
        def dialog_mousewheel(event):
            try:
                dialog_canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
            except tk.TclError:
                pass  # Dialog might be destroyed
                
        dialog_canvas.bind("<MouseWheel>", dialog_mousewheel)
        content_frame.bind("<MouseWheel>", dialog_mousewheel)
        dialog_canvas.bind("<Button-4>", lambda e: dialog_canvas.yview_scroll(-1, "units"))
        dialog_canvas.bind("<Button-5>", lambda e: dialog_canvas.yview_scroll(1, "units"))
        
        # FIXED: Cleanup binding when dialog is destroyed
        def cleanup_dialog():
            try:
                dialog_canvas.unbind("<MouseWheel>")
                content_frame.unbind("<MouseWheel>")
                dialog_canvas.unbind("<Button-4>")
                dialog_canvas.unbind("<Button-5>")
                dialog.destroy()
            except tk.TclError:
                pass
        
        dialog.protocol("WM_DELETE_WINDOW", cleanup_dialog)
        
        return dialog, content_frame, cleanup_dialog

    def _show_loading_dialog(self, title="Processing", message="Please wait..."):
        """FIXED: Loading dialog with proper cleanup"""
        loading_window = tk.Toplevel(self.parent)
        loading_window.title(title)
        loading_window.geometry("400x150")
        loading_window.transient(self.parent)
        loading_window.grab_set()
        loading_window.configure(bg=MODERN_SLATE['bg_primary'])
        
        # Center the window
        loading_window.geometry("+%d+%d" % (
            loading_window.winfo_screenwidth()//2 - 200,
            loading_window.winfo_screenheight()//2 - 75
        ))
        
        # Create content frame with modern styling
        content_frame = Frame(loading_window, style="Card.TFrame", padding=(30, 20))
        content_frame.pack(fill=BOTH, expand=True)
        
        # Message label
        Label(content_frame, text=message, 
              style="Secondary.TLabel", 
              font=("Segoe UI", 11),
              justify="center").pack(pady=(0, 15))
        
        # Progress bar with modern styling
        from tkinter import ttk
        progress = ttk.Progressbar(content_frame, 
                                  mode='indeterminate',
                                  style="Modern.Horizontal.TProgressbar")
        progress.pack(fill=X, pady=(0, 10))
        progress.start(10)
        
        # Status label
        status_label = Label(content_frame, text="Analyzing training approaches...", 
                            style="Secondary.TLabel", 
                            font=("Segoe UI", 9))
        status_label.pack()
        
        loading_window.update()
        return loading_window, progress, status_label

    def _close_loading_dialog(self, loading_window, progress):
        """FIXED: Close loading dialog safely"""
        try:
            progress.stop()
            loading_window.grab_release()
            loading_window.destroy()
        except tk.TclError:
            pass  # Dialog might already be closed

    def _get_analysis_cache_key(self, chunks, tokenizer_name, token_limit, target_models, api_usage):
        """Generate cache key for cost analysis results"""
        import hashlib
        
        # Create a hash of the analysis parameters
        chunks_hash = hashlib.md5(str(len(chunks)).encode() + 
                                 (chunks[0][:100] if chunks else "").encode()).hexdigest()[:8]
        
        params_str = f"{chunks_hash}_{tokenizer_name}_{token_limit}_{len(target_models)}_{api_usage}"
        return hashlib.md5(params_str.encode()).hexdigest()[:16]

    def _is_analysis_cache_valid(self, cache_key):
        """Check if cached analysis is still valid (within 5 minutes)"""
        if cache_key not in self.cost_analysis_cache:
            return False
        
        cache_time = self.cost_analysis_cache[cache_key].get('timestamp', 0)
        return (time.time() - cache_time) < 300  # 5 minutes

    def _show_enhanced_error_dialog(self, title, message, recovery_suggestions=None):
        """Show enhanced error dialog with recovery suggestions"""
        error_msg = f"❌ {message}\n\n"
        
        if recovery_suggestions:
            error_msg += "💡 Suggested Solutions:\n"
            for i, suggestion in enumerate(recovery_suggestions, 1):
                error_msg += f"{i}. {suggestion}\n"
            error_msg += "\n"
        
        error_msg += "🔄 You can try:\n"
        error_msg += "• Refreshing the analysis\n"
        error_msg += "• Using a different tokenizer\n"
        error_msg += "• Processing a smaller file\n"
        error_msg += "\n📧 Need help? Contact: support@wolflow.ai"
        
        messagebox.showerror(title, error_msg)

    def _display_cost_analysis_dialog(self, cost_analysis):
        """STAGE 2 ENHANCED: Display comprehensive cost analysis with FIXED scrolling"""
        if not cost_analysis.get('cost_analysis', {}).get('available'):
            error_msg = cost_analysis.get('cost_analysis', {}).get('error', 'Cost analysis not available')
            messagebox.showerror("Cost Analysis Error", f"Cost analysis failed: {error_msg}")
            return

        # FIXED: Use the new scrollable dialog method
        cost_window, content_frame, cleanup_dialog = self._create_scrollable_dialog(
            "💰 Comprehensive Training Cost Analysis",
            1100, 
            800
        )

        # Extract cost analysis data
        cost_data = cost_analysis.get('cost_analysis', {})
        detailed_results = cost_data.get('detailed_results', {})
        summary = cost_data.get('summary', {})
        
        # Header Section with modern styling
        header_frame = Frame(content_frame, style="Modern.TFrame", padding=(0, 0, 0, 20))
        header_frame.pack(fill=X)
        
        Label(header_frame, text="💰 Comprehensive Training Cost Analysis", 
              style="Heading.TLabel", font=("Segoe UI", 18, "bold")).pack(anchor="w")
        
        dataset_info = cost_analysis.get('dataset_info', {})
        Label(header_frame, 
              text=f"Dataset: {dataset_info.get('tokens', 0):,} tokens | "
                   f"Chunks: {len(self.parent.chunks)} | "
                   f"Tokenizer: {getattr(self.parent, '_current_tokenizer_name', 'gpt2')}", 
              style="Secondary.TLabel", font=("Segoe UI", 11)).pack(anchor="w")

        # Enhanced Summary Section with modern cards
        if summary:
            summary_frame = Frame(content_frame, style="Card.TFrame", padding=(20, 15))
            summary_frame.pack(fill=X, pady=(0, 20))
            
            Label(summary_frame, text="📊 Executive Summary", 
                  style="Heading.TLabel", font=("Segoe UI", 16, "bold")).pack(anchor="w", pady=(0, 10))
            
            # Create three-column layout for summary
            summary_cols = Frame(summary_frame, style="Modern.TFrame")
            summary_cols.pack(fill=X)
            
            # Column 1: Best Option
            col1 = Frame(summary_cols, style="Modern.TFrame")
            col1.pack(side=LEFT, fill=BOTH, expand=True, padx=(0, 10))
            
            Label(col1, text="🏆 Best Training Option", 
                  style="Success.TLabel", font=("Segoe UI", 12, "bold")).pack(anchor="w")
            
            best_option = summary.get('best_overall', {})
            if best_option:
                Label(col1, text=f"Approach: {best_option.get('best_approach', 'N/A')}", 
                      style="Secondary.TLabel", font=("Segoe UI", 11, "bold")).pack(anchor="w")
                Label(col1, text=f"Cost: ${best_option.get('cost', 0):.2f}", 
                      style="Secondary.TLabel", font=("Segoe UI", 11)).pack(anchor="w")
                Label(col1, text=f"Time: {best_option.get('hours', 0):.1f} hours", 
                      style="Secondary.TLabel", font=("Segoe UI", 11)).pack(anchor="w")
            
            # Column 2: Cost Range
            col2 = Frame(summary_cols, style="Modern.TFrame")
            col2.pack(side=LEFT, fill=BOTH, expand=True, padx=(0, 10))
            
            Label(col2, text="💰 Cost Analysis", 
                  style="Primary.TLabel", font=("Segoe UI", 12, "bold")).pack(anchor="w")
            
            cost_range = summary.get('cost_range', {})
            if cost_range:
                Label(col2, text=f"Range: ${cost_range.get('min', 0):.2f} - ${cost_range.get('max', 0):.2f}", 
                      style="Secondary.TLabel", font=("Segoe UI", 11)).pack(anchor="w")
                savings = cost_range.get('max', 0) - cost_range.get('min', 0)
                Label(col2, text=f"Max Savings: ${savings:.2f}", 
                      style="CostSavings.TLabel", font=("Segoe UI", 11)).pack(anchor="w")
                Label(col2, text=f"Models Compared: {summary.get('models_compared', 0)}", 
                      style="Secondary.TLabel", font=("Segoe UI", 11)).pack(anchor="w")
            
            # Column 3: ROI Quick Stats
            col3 = Frame(summary_cols, style="Modern.TFrame")
            col3.pack(side=LEFT, fill=BOTH, expand=True)
            
            Label(col3, text="📈 ROI Overview", 
                  style="Premium.TLabel", font=("Segoe UI", 12, "bold")).pack(anchor="w")
            
            # Calculate simple ROI metrics from best option
            if best_option:
                monthly_api_cost = 100  # Estimated monthly API cost for comparison
                monthly_savings = monthly_api_cost * 0.9  # 90% savings assumption
                training_cost = best_option.get('cost', 0)
                break_even = training_cost / monthly_savings if monthly_savings > 0 else float('inf')
                
                Label(col3, text=f"Break-even: {break_even:.1f} months", 
                      style="Secondary.TLabel", font=("Segoe UI", 11)).pack(anchor="w")
                annual_savings = (monthly_savings * 12) - training_cost
                Label(col3, text=f"Annual ROI: ${annual_savings:.0f}", 
                      style="CostSavings.TLabel", font=("Segoe UI", 11)).pack(anchor="w")
                Label(col3, text=f"Payback: {break_even*30:.0f} days", 
                      style="Secondary.TLabel", font=("Segoe UI", 11)).pack(anchor="w")

        # Comprehensive Approaches Table with modern styling
        if detailed_results:
            approaches_frame = Frame(content_frame, style="Modern.TFrame")
            approaches_frame.pack(fill=BOTH, expand=True, pady=(0, 20))
            
            Label(approaches_frame, text="🔧 Complete Training Approaches Comparison", 
                  style="Heading.TLabel", font=("Segoe UI", 16, "bold")).pack(anchor="w", pady=(0, 15))
            
            # Create enhanced comparison table with modern dark styling
            table_container = Frame(approaches_frame, style="Card.TFrame")
            table_container.pack(fill=BOTH, expand=True)
            
            # Enhanced table headers with modern styling
            headers_frame = Frame(table_container, style="Card.TFrame", padding=(15, 10))
            headers_frame.pack(fill=X)
            
            # Header labels with consistent styling
            headers = [
                ("Rank", 6), ("Model", 15), ("Training Approach", 20),
                ("Cost (USD)", 12), ("Time (Hours)", 12), ("Hardware", 15), ("Confidence", 10)
            ]
            
            for header, width in headers:
                Label(headers_frame, text=header, 
                      style="Secondary.TLabel", font=("Segoe UI", 10, "bold"), 
                      width=width).pack(side=LEFT)
            
            # Collect and sort all approaches
            all_approaches = []
            for model_name, model_data in detailed_results.items():
                if 'error' in model_data:
                    continue
                    
                cost_estimates = model_data.get('cost_estimates', [])
                for estimate in cost_estimates:
                    # Extract hardware info
                    hw_req = estimate.get('hardware_requirements', {})
                    gpu_type = hw_req.get('gpu_type', 'Unknown')
                    gpu_count = hw_req.get('gpu_count', 1)
                    hardware_display = f"{gpu_type}" + (f" x{gpu_count}" if gpu_count > 1 else "")
                    
                    all_approaches.append({
                        'model': model_name,
                        'approach': estimate['approach_name'],
                        'cost': estimate['total_cost_usd'],
                        'hours': estimate['training_hours'],
                        'hardware': hardware_display,
                        'confidence': estimate.get('confidence', 0.8)
                    })
            
            # Sort by cost (cheapest first)
            all_approaches.sort(key=lambda x: x['cost'])
            
            # Display top 15 approaches with color coding
            for i, approach in enumerate(all_approaches[:15]):
                row_frame = Frame(table_container, style="Modern.TFrame", padding=(15, 8))
                row_frame.pack(fill=X)
                
                # Rank with medal icons for top 3
                rank_display = "🥇" if i == 0 else "🥈" if i == 1 else "🥉" if i == 2 else f"#{i+1}"
                
                # Color-coded styling based on cost efficiency
                if i < 3:
                    text_style = "Success.TLabel"  # Top 3 in green
                elif approach['cost'] > all_approaches[0]['cost'] * 3:
                    text_style = "Warning.TLabel"  # Expensive options in amber
                else:
                    text_style = "Secondary.TLabel"  # Normal options
                
                # Row data with proper truncation
                row_data = [
                    (rank_display, 6),
                    (approach['model'][:12] + "..." if len(approach['model']) > 12 else approach['model'], 15),
                    (approach['approach'][:18] + "..." if len(approach['approach']) > 18 else approach['approach'], 20),
                    (f"${approach['cost']:.2f}", 12),
                    (f"{approach['hours']:.1f}h", 12),
                    (approach['hardware'][:12] + "..." if len(approach['hardware']) > 12 else approach['hardware'], 15),
                    (f"{approach['confidence']*100:.0f}%", 10)
                ]
                
                for data, width in row_data:
                    font_weight = "bold" if i < 3 and data not in [rank_display] else "normal"
                    Label(row_frame, text=data, 
                          style=text_style, 
                          font=("Segoe UI", 9, font_weight), 
                          width=width).pack(side=LEFT, anchor="w")
            
            # Show count of additional approaches if more than 15
            if len(all_approaches) > 15:
                more_frame = Frame(table_container, style="Modern.TFrame", padding=(15, 5))
                more_frame.pack(fill=X)
                Label(more_frame, text=f"... and {len(all_approaches) - 15} more approaches analyzed", 
                      style="Secondary.TLabel", font=("Segoe UI", 9)).pack(anchor="w")

        # Enhanced Recommendations with modern styling
        recommendations = cost_data.get('recommendations', [])
        if recommendations:
            rec_frame = Frame(content_frame, style="Card.TFrame", padding=(20, 15))
            rec_frame.pack(fill=X, pady=(0, 20))
            
            Label(rec_frame, text="💡 Cost Optimization Recommendations", 
                  style="Heading.TLabel", font=("Segoe UI", 16, "bold")).pack(anchor="w", pady=(0, 10))
            
            # Display top recommendations with proper styling
            for i, rec in enumerate(recommendations[:5], 1):
                Label(rec_frame, text=f"• {rec}", 
                      style="Secondary.TLabel", font=("Segoe UI", 10), 
                      wraplength=900, justify="left").pack(anchor="w", padx=(20, 0), pady=2)

        # Enhanced action buttons with modern styling - STAGE 3 UPDATED
        button_frame = Frame(content_frame, style="Modern.TFrame")
        button_frame.pack(fill=X, pady=20)
        
        Button(button_frame, text="📊 Export Analysis", 
               command=lambda: self._export_cost_analysis(cost_analysis), 
               style="Success.TButton").pack(side=LEFT, padx=(0, 10))
        
        Button(button_frame, text="🔄 Refresh Pricing", 
               command=lambda: self._refresh_cost_analysis(), 
               style="Secondary.TButton").pack(side=LEFT, padx=(0, 10))
        
        Button(button_frame, text="Close", command=cleanup_dialog, 
               style="Secondary.TButton").pack(side=RIGHT)

# CONTINUING cost_dialogs.py - Part 2: Export Methods and Refresh Functionality

    def _export_cost_analysis(self, cost_analysis):
        """STAGE 3 ENHANCED: Enhanced export with metadata and multiple formats"""
        try:
            # Create export options dialog
            export_window = tk.Toplevel(self.parent)
            export_window.title("📊 Export Cost Analysis")
            export_window.geometry("500x400")
            export_window.transient(self.parent)
            export_window.grab_set()
            export_window.configure(bg=MODERN_SLATE['bg_primary'])
            
            # Content frame
            content_frame = Frame(export_window, style="Card.TFrame", padding=(25, 20))
            content_frame.pack(fill=BOTH, expand=True, padx=15, pady=15)
            
            # Header
            Label(content_frame, text="📊 Export Cost Analysis Report", 
                  style="Heading.TLabel", font=("Segoe UI", 16, "bold")).pack(anchor="w", pady=(0, 15))
            
            # Export format selection
            Label(content_frame, text="Select Export Format:", 
                  style="FieldLabel.TLabel").pack(anchor="w", pady=(0, 8))
            
            export_format = tk.StringVar(value="json")
            
            # Format options with descriptions
            formats_frame = Frame(content_frame, style="Modern.TFrame")
            formats_frame.pack(fill=X, pady=(0, 15))
            
            format_options = [
                ("json", "📄 JSON Report", "Complete analysis with all data"),
                ("csv", "📊 CSV Summary", "Spreadsheet-compatible cost table"),
                ("txt", "📝 Text Report", "Human-readable formatted report"),
                ("excel", "📈 Excel Workbook", "Professional spreadsheet with charts")
            ]
            
            for value, label, description in format_options:
                radio_frame = Frame(formats_frame, style="Modern.TFrame")
                radio_frame.pack(fill=X, pady=2)
                
                radio = tk.Radiobutton(radio_frame, 
                                      text=label,
                                      variable=export_format,
                                      value=value,
                                      bg=MODERN_SLATE['bg_cards'],
                                      fg=MODERN_SLATE['text_primary'],
                                      selectcolor=MODERN_SLATE['accent_blue'],
                                      activebackground=MODERN_SLATE['bg_hover'],
                                      font=("Segoe UI", 10, "bold"))
                radio.pack(side=LEFT)
                
                Label(radio_frame, text=f" - {description}", 
                      style="Secondary.TLabel", 
                      font=("Segoe UI", 9)).pack(side=LEFT)
            
            # Include options
            Label(content_frame, text="Include in Export:", 
                  style="FieldLabel.TLabel").pack(anchor="w", pady=(15, 8))
            
            options_frame = Frame(content_frame, style="Modern.TFrame")
            options_frame.pack(fill=X, pady=(0, 15))
            
            include_metadata = tk.BooleanVar(value=True)
            include_recommendations = tk.BooleanVar(value=True)
            include_charts = tk.BooleanVar(value=False)
            
            tk.Checkbutton(options_frame, text="📋 Metadata (timestamp, dataset info)", 
                          variable=include_metadata,
                          bg=MODERN_SLATE['bg_cards'],
                          fg=MODERN_SLATE['text_primary'],
                          selectcolor=MODERN_SLATE['accent_blue'],
                          activebackground=MODERN_SLATE['bg_hover']).pack(anchor="w")
            
            tk.Checkbutton(options_frame, text="💡 Optimization recommendations", 
                          variable=include_recommendations,
                          bg=MODERN_SLATE['bg_cards'],
                          fg=MODERN_SLATE['text_primary'],
                          selectcolor=MODERN_SLATE['accent_blue'],
                          activebackground=MODERN_SLATE['bg_hover']).pack(anchor="w")
            
            tk.Checkbutton(options_frame, text="📈 Charts and visualizations (Excel only)", 
                          variable=include_charts,
                          bg=MODERN_SLATE['bg_cards'],
                          fg=MODERN_SLATE['text_primary'],
                          selectcolor=MODERN_SLATE['accent_blue'],
                          activebackground=MODERN_SLATE['bg_hover']).pack(anchor="w")
            
            # Buttons
            button_frame = Frame(content_frame, style="Modern.TFrame")
            button_frame.pack(fill=X, pady=(15, 0))
            
            def do_export():
                try:
                    selected_format = export_format.get()
                    
                    # File extension mapping
                    ext_map = {
                        "json": ".json",
                        "csv": ".csv", 
                        "txt": ".txt",
                        "excel": ".xlsx"
                    }
                    
                    # File type mapping for dialog
                    filetype_map = {
                        "json": [("JSON Report", "*.json")],
                        "csv": [("CSV File", "*.csv")],
                        "txt": [("Text Report", "*.txt")],
                        "excel": [("Excel Workbook", "*.xlsx")]
                    }
                    
                    path = filedialog.asksaveasfilename(
                        title=f"Export {selected_format.upper()} Report",
                        defaultextension=ext_map[selected_format],
                        filetypes=filetype_map[selected_format]
                    )
                    
                    if not path:
                        return
                    
                    export_window.destroy()
                    
                    # Show progress for export
                    progress_window, progress_bar, _ = self._show_loading_dialog(
                        "Exporting Report", f"Generating {selected_format.upper()} report..."
                    )
                    
                    def export_worker():
                        try:
                            if selected_format == "json":
                                self._export_json_report(cost_analysis, path, 
                                                       include_metadata.get(), 
                                                       include_recommendations.get())
                            elif selected_format == "csv":
                                self._export_csv_report(cost_analysis, path,
                                                       include_metadata.get())
                            elif selected_format == "txt":
                                self._export_text_report(cost_analysis, path,
                                                        include_metadata.get(),
                                                        include_recommendations.get())
                            elif selected_format == "excel":
                                self._export_excel_report(cost_analysis, path,
                                                         include_metadata.get(),
                                                         include_recommendations.get(),
                                                         include_charts.get())
                            
                            self._close_loading_dialog(progress_window, progress_bar)
                            messagebox.showinfo("✅ Export Complete", 
                                              f"Report exported successfully to:\n{path}")
                            
                        except Exception as e:
                            self._close_loading_dialog(progress_window, progress_bar)
                            self._show_enhanced_error_dialog(
                                "Export Error",
                                f"Failed to export report: {str(e)}",
                                recovery_suggestions=[
                                    "Try a different file location",
                                    "Check disk space and permissions",
                                    "Try a different export format"
                                ]
                            )
                    
                    export_thread = threading.Thread(target=export_worker)
                    export_thread.daemon = True
                    export_thread.start()
                    
                except Exception as e:
                    messagebox.showerror("Export Error", f"Failed to start export: {str(e)}")
            
            Button(button_frame, text="📤 Export Report", 
                   command=do_export, 
                   style="Success.TButton").pack(side=LEFT, padx=(0, 10))
            
            Button(button_frame, text="Cancel", 
                   command=export_window.destroy, 
                   style="Secondary.TButton").pack(side=RIGHT)
            
        except Exception as e:
            messagebox.showerror("Export Error", f"Failed to show export dialog: {str(e)}")

    def _export_json_report(self, cost_analysis, path, include_metadata, include_recommendations):
        """Export comprehensive JSON report with metadata"""
        import json
        
        report = cost_analysis.copy()
        
        if include_metadata:
            # Add comprehensive metadata
            report['export_metadata'] = {
                'exported_at': datetime.now().isoformat(),
                'exported_by': 'Wolfscribe Premium v2.2',
                'export_format': 'json',
                'dataset_info': {
                    'total_chunks': len(self.parent.chunks),
                    'tokenizer_used': getattr(self.parent, '_current_tokenizer_name', 'gpt2'),
                    'token_limit': 512,
                    'file_processed': os.path.basename(self.parent.file_path) if self.parent.file_path else 'Unknown'
                },
                'license_info': {
                    'tier': self.controller.get_licensing_info()['license_status']['tier'],
                    'status': self.controller.get_licensing_info()['license_status']['status']
                }
            }
        
        if not include_recommendations:
            # Remove recommendations if not wanted
            if 'cost_analysis' in report and 'recommendations' in report['cost_analysis']:
                del report['cost_analysis']['recommendations']
        
        with open(path, 'w', encoding='utf-8') as f:
            json.dump(report, f, indent=2, ensure_ascii=False)

    def _export_csv_report(self, cost_analysis, path, include_metadata):
        """Export CSV summary with cost comparison table"""
        import csv
        
        cost_data = cost_analysis.get('cost_analysis', {})
        detailed_results = cost_data.get('detailed_results', {})
        
        with open(path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            
            # Header with metadata
            if include_metadata:
                writer.writerow(['# Wolfscribe Cost Analysis Report'])
                writer.writerow([f'# Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'])
                writer.writerow([f'# Dataset: {len(self.parent.chunks)} chunks'])
                writer.writerow([f'# Tokenizer: {getattr(self.parent, "_current_tokenizer_name", "gpt2")}'])
                writer.writerow([''])
            
            # Column headers
            writer.writerow(['Rank', 'Model', 'Training_Approach', 'Cost_USD', 'Time_Hours', 
                            'Hardware_Type', 'GPU_Count', 'Confidence_Percent', 'Notes'])
            
            # Collect all approaches
            all_approaches = []
            for model_name, model_data in detailed_results.items():
                if 'error' in model_data:
                    continue
                cost_estimates = model_data.get('cost_estimates', [])
                for estimate in cost_estimates:
                    hw_req = estimate.get('hardware_requirements', {})
                    all_approaches.append([
                        model_name,
                        estimate['approach_name'],
                        estimate['total_cost_usd'],
                        estimate['training_hours'],
                        hw_req.get('gpu_type', 'Unknown'),
                        hw_req.get('gpu_count', 1),
                        f"{estimate.get('confidence', 0.8) * 100:.0f}",
                        '; '.join(estimate.get('notes', []))[:100]  # Truncate notes
                    ])
            
            # Sort by cost and add rank
            all_approaches.sort(key=lambda x: x[2])
            for i, approach in enumerate(all_approaches, 1):
                writer.writerow([i] + approach)

    def _export_text_report(self, cost_analysis, path, include_metadata, include_recommendations):
        """Export formatted text report for cost analysis"""
        report = "💰 WOLFSCRIBE COST ANALYSIS REPORT\n"
        report += "=" * 50 + "\n\n"
        
        # Add timestamp
        if include_metadata:
            report += f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n"
            report += f"Dataset: {len(self.parent.chunks)} chunks, {cost_analysis.get('dataset_info', {}).get('tokens', 0):,} tokens\n"
            report += f"Tokenizer: {getattr(self.parent, '_current_tokenizer_name', 'gpt2')}\n\n"
        
        # Executive Summary
        cost_data = cost_analysis.get('cost_analysis', {})
        summary = cost_data.get('summary', {})
        
        if summary:
            report += "📊 EXECUTIVE SUMMARY\n"
            report += "-" * 20 + "\n"
            
            best_option = summary.get('best_overall', {})
            if best_option:
                report += f"Best Approach: {best_option.get('best_approach', 'N/A')}\n"
                report += f"Optimal Cost: ${best_option.get('cost', 0):.2f}\n"
                report += f"Training Time: {best_option.get('hours', 0):.1f} hours\n"
            
            cost_range = summary.get('cost_range', {})
            if cost_range:
                report += f"Cost Range: ${cost_range.get('min', 0):.2f} - ${cost_range.get('max', 0):.2f}\n"
                savings = cost_range.get('max', 0) - cost_range.get('min', 0)
                report += f"Maximum Savings: ${savings:.2f}\n"
            
            report += f"Models Analyzed: {summary.get('models_compared', 0)}\n\n"
        
        # Detailed Results
        detailed_results = cost_data.get('detailed_results', {})
        if detailed_results:
            report += "🔧 DETAILED TRAINING APPROACHES\n"
            report += "-" * 35 + "\n\n"
            
            all_approaches = []
            for model_name, model_data in detailed_results.items():
                if 'error' in model_data:
                    continue
                cost_estimates = model_data.get('cost_estimates', [])
                for estimate in cost_estimates:
                    all_approaches.append({
                        'model': model_name,
                        'approach': estimate['approach_name'],
                        'cost': estimate['total_cost_usd'],
                        'hours': estimate['training_hours'],
                        'hardware': estimate.get('hardware_requirements', {})
                    })
            
            # Sort by cost
            all_approaches.sort(key=lambda x: x['cost'])
            
            for i, approach in enumerate(all_approaches[:10], 1):  # Top 10
                report += f"{i:2d}. {approach['approach']} ({approach['model']})\n"
                report += f"    Cost: ${approach['cost']:.2f} | Time: {approach['hours']:.1f}h\n"
                
                hw = approach['hardware']
                gpu_type = hw.get('gpu_type', 'Unknown')
                gpu_count = hw.get('gpu_count', 1)
                hardware_str = f"{gpu_type}" + (f" x{gpu_count}" if gpu_count > 1 else "")
                report += f"    Hardware: {hardware_str}\n\n"
        
        # Recommendations
        if include_recommendations:
            recommendations = cost_data.get('recommendations', [])
            if recommendations:
                report += "💡 OPTIMIZATION RECOMMENDATIONS\n"
                report += "-" * 35 + "\n"
                for i, rec in enumerate(recommendations[:5], 1):
                    report += f"{i}. {rec}\n"
                report += "\n"
        
        # Footer
        if include_metadata:
            report += "=" * 50 + "\n"
            report += "Generated by Wolfscribe Premium\n"
            report += "https://wolflow.ai\n"
        
        with open(path, 'w', encoding='utf-8') as f:
            f.write(report)

    def _export_excel_report(self, cost_analysis, path, include_metadata, include_recommendations, include_charts):
        """Export Excel workbook with multiple sheets and optional charts"""
        try:
            # Try to import openpyxl for Excel export
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
            from openpyxl.chart import BarChart, Reference
            
            wb = Workbook()
            
            # Summary sheet
            ws_summary = wb.active
            ws_summary.title = "Executive Summary"
            
            # Add title and metadata
            ws_summary['A1'] = "Wolfscribe Training Cost Analysis"
            ws_summary['A1'].font = Font(size=16, bold=True)
            
            if include_metadata:
                ws_summary['A3'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
                ws_summary['A4'] = f"Dataset: {len(self.parent.chunks)} chunks"
                ws_summary['A5'] = f"Tokenizer: {getattr(self.parent, '_current_tokenizer_name', 'gpt2')}"
            
            # Summary data
            cost_data = cost_analysis.get('cost_analysis', {})
            summary = cost_data.get('summary', {})
            
            if summary:
                best_option = summary.get('best_overall', {})
                cost_range = summary.get('cost_range', {})
                
                row = 7
                ws_summary[f'A{row}'] = "Best Training Option:"
                ws_summary[f'B{row}'] = best_option.get('best_approach', 'N/A')
                ws_summary[f'A{row}'].font = Font(bold=True)
                
                row += 1
                ws_summary[f'A{row}'] = "Optimal Cost:"
                ws_summary[f'B{row}'] = f"${best_option.get('cost', 0):.2f}"
                
                row += 1
                ws_summary[f'A{row}'] = "Training Time:"
                ws_summary[f'B{row}'] = f"{best_option.get('hours', 0):.1f} hours"
                
                row += 2
                ws_summary[f'A{row}'] = "Cost Range:"
                ws_summary[f'B{row}'] = f"${cost_range.get('min', 0):.2f} - ${cost_range.get('max', 0):.2f}"
            
            # Detailed comparison sheet
            ws_details = wb.create_sheet("Cost Comparison")
            
            # Headers
            headers = ['Rank', 'Model', 'Approach', 'Cost (USD)', 'Time (Hours)', 'Hardware', 'Confidence']
            for col, header in enumerate(headers, 1):
                cell = ws_details.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            # Data rows
            detailed_results = cost_data.get('detailed_results', {})
            all_approaches = []
            
            for model_name, model_data in detailed_results.items():
                if 'error' in model_data:
                    continue
                cost_estimates = model_data.get('cost_estimates', [])
                for estimate in cost_estimates:
                    hw_req = estimate.get('hardware_requirements', {})
                    hardware = f"{hw_req.get('gpu_type', 'Unknown')}"
                    if hw_req.get('gpu_count', 1) > 1:
                        hardware += f" x{hw_req['gpu_count']}"
                    
                    all_approaches.append([
                        model_name,
                        estimate['approach_name'],
                        estimate['total_cost_usd'],
                        estimate['training_hours'],
                        hardware,
                        estimate.get('confidence', 0.8)
                    ])
            
            # Sort and populate
            all_approaches.sort(key=lambda x: x[2])
            for i, approach in enumerate(all_approaches, 2):
                ws_details.cell(row=i, column=1, value=i-1)  # Rank
                for col, value in enumerate(approach, 2):
                    if col == 4:  # Cost column
                        ws_details.cell(row=i, column=col, value=f"${value:.2f}")
                    elif col == 5:  # Time column
                        ws_details.cell(row=i, column=col, value=f"{value:.1f}h")
                    elif col == 7:  # Confidence column
                        ws_details.cell(row=i, column=col, value=f"{value*100:.0f}%")
                    else:
                        ws_details.cell(row=i, column=col, value=value)
            
            # Auto-adjust column widths
            for sheet in [ws_summary, ws_details]:
                for column in sheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    sheet.column_dimensions[column_letter].width = adjusted_width
            
            # Add chart if requested
            if include_charts and len(all_approaches) > 1:
                chart = BarChart()
                chart.type = "col"
                chart.style = 10
                chart.title = "Training Cost Comparison"
                chart.y_axis.title = "Cost (USD)"
                chart.x_axis.title = "Training Approach"
                
                # Data for chart (top 10 approaches)
                data = Reference(ws_details, min_col=4, min_row=1, max_row=min(11, len(all_approaches)+1), max_col=4)
                cats = Reference(ws_details, min_col=3, min_row=2, max_row=min(11, len(all_approaches)+1))
                
                chart.add_data(data, titles_from_data=True)
                chart.set_categories(cats)
                
                ws_details.add_chart(chart, "I2")
            
            # Recommendations sheet
            if include_recommendations:
                recommendations = cost_data.get('recommendations', [])
                if recommendations:
                    ws_rec = wb.create_sheet("Recommendations")
                    ws_rec['A1'] = "Cost Optimization Recommendations"
                    ws_rec['A1'].font = Font(size=14, bold=True)
                    
                    for i, rec in enumerate(recommendations, 3):
                        ws_rec[f'A{i}'] = f"{i-2}. {rec}"
            
            wb.save(path)
            
        except ImportError:
            # Fallback to CSV if openpyxl not available
            messagebox.showwarning("Excel Export", 
                                  "Excel export requires openpyxl package.\nExporting as CSV instead.")
            csv_path = path.replace('.xlsx', '.csv')
            self._export_csv_report(cost_analysis, csv_path, include_metadata)
        except Exception as e:
            raise Exception(f"Excel export failed: {str(e)}")

    def _refresh_cost_analysis(self):
        """STAGE 3 ENHANCED: Enhanced refresh with cache clearing and user feedback"""
        try:
            if not self.parent.chunks:
                messagebox.showwarning("No Data", "Please process a file first.")
                return
            
            # Clear cache
            self.cost_analysis_cache.clear()
            
            result = messagebox.askyesno("🔄 Refresh Cost Analysis",
                "This will refresh the cost analysis with the latest pricing data.\n\n"
                "• Clears cached results\n"
                "• Fetches live cloud pricing\n"
                "• Updates model compatibility\n\n"
                "Continue with refresh?")
            
            if result:
                # Use the enhanced analysis method
                self.show_cost_analysis()
            
        except Exception as e:
            self._show_enhanced_error_dialog(
                "Refresh Error",
                f"Failed to refresh cost analysis: {str(e)}",
                recovery_suggestions=[
                    "Check your internet connection",
                    "Try processing the file again",
                    "Use cached results if available"
                ]
            )
# wolfscribe/processing/extractors/xlsx_extractor.py
"""
Excel XLSX file extractor for Wolfscribe

Handles .xlsx and .xls files by extracting text content from multiple sheets,
intelligently identifying text vs data content, and handling cell comments.
"""

import os
from typing import List, Dict, Any, Tuple


def extract_text(path: str) -> str:
    """
    Extract text from Excel XLSX/XLS files
    
    Args:
        path (str): Path to the Excel file
        
    Returns:
        str: Extracted text content from all sheets
        
    Raises:
        RuntimeError: If Excel extraction fails
    """
    if not os.path.exists(path):
        raise RuntimeError(f"Excel file not found: {path}")
    
    file_ext = os.path.splitext(path)[1].lower()
    
    try:
        if file_ext in ['.xlsx', '.xlsm']:
            return _extract_xlsx(path)
        elif file_ext == '.xls':
            return _extract_xls(path)
        else:
            raise RuntimeError(f"Unsupported Excel format: {file_ext}")
            
    except Exception as e:
        if "Excel" in str(e):
            raise
        else:
            raise RuntimeError(f"Excel extraction failed: {str(e)}")


def _extract_xlsx(path: str) -> str:
    """Extract text from XLSX files using openpyxl"""
    try:
        import openpyxl
        from openpyxl.utils.exceptions import InvalidFileException
    except ImportError:
        # Fallback to pandas
        return _extract_with_pandas(path)
    
    try:
        # Load workbook
        workbook = openpyxl.load_workbook(path, data_only=True)
        
        all_text = []
        
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            
            # Extract text from this sheet
            sheet_text = _extract_sheet_text_openpyxl(sheet, sheet_name)
            if sheet_text:
                all_text.append(sheet_text)
        
        workbook.close()
        
        if not all_text:
            return ""
        
        return "\n\n".join(all_text)
        
    except InvalidFileException:
        raise RuntimeError("Excel file is corrupted or password protected")
    except Exception as e:
        # Fallback to pandas
        try:
            return _extract_with_pandas(path)
        except Exception:
            raise RuntimeError(f"Cannot read Excel file: {str(e)}")


def _extract_xls(path: str) -> str:
    """Extract text from XLS files using pandas (requires xlrd)"""
    return _extract_with_pandas(path)


def _extract_with_pandas(path: str) -> str:
    """Extract text using pandas (fallback method)"""
    try:
        import pandas as pd
    except ImportError:
        raise RuntimeError(
            "Excel extraction requires pandas library. "
            "Install with: pip install pandas openpyxl"
        )
    
    try:
        # Get all sheet names
        xl_file = pd.ExcelFile(path)
        sheet_names = xl_file.sheet_names
        
        all_text = []
        
        for sheet_name in sheet_names:
            try:
                # Read the sheet
                df = pd.read_excel(path, sheet_name=sheet_name, dtype=str, na_filter=False)
                
                # Extract text from this sheet
                sheet_text = _extract_sheet_text_pandas(df, sheet_name)
                if sheet_text:
                    all_text.append(sheet_text)
                    
            except Exception as e:
                # Skip problematic sheets but continue with others
                print(f"Warning: Could not process sheet '{sheet_name}': {str(e)}")
                continue
        
        xl_file.close()
        
        if not all_text:
            return ""
        
        return "\n\n".join(all_text)
        
    except Exception as e:
        error_msg = str(e).lower()
        if "password" in error_msg or "encrypted" in error_msg:
            raise RuntimeError("Excel file is password protected")
        elif "corrupted" in error_msg or "invalid" in error_msg:
            raise RuntimeError("Excel file is corrupted or invalid")
        else:
            raise RuntimeError(f"Cannot read Excel file: {str(e)}")


def _extract_sheet_text_openpyxl(sheet, sheet_name: str) -> str:
    """Extract text from a sheet using openpyxl"""
    sheet_content = []
    
    # Add sheet name as header
    if sheet_name and sheet_name.lower() not in ['sheet1', 'sheet', 'data']:
        sheet_content.append(f"=== {sheet_name} ===")
    
    # Get sheet dimensions
    if sheet.max_row is None or sheet.max_column is None:
        return ""  # Empty sheet
    
    # Analyze the sheet to identify text vs data regions
    text_analysis = _analyze_sheet_content_openpyxl(sheet)
    
    # Extract text based on analysis
    if text_analysis['has_headers']:
        sheet_content.extend(_extract_with_headers_openpyxl(sheet, text_analysis))
    else:
        sheet_content.extend(_extract_without_headers_openpyxl(sheet, text_analysis))
    
    # Extract cell comments
    comments = _extract_comments_openpyxl(sheet)
    if comments:
        sheet_content.append("--- Comments ---")
        sheet_content.extend(comments)
    
    if len(sheet_content) <= 1:  # Only header or empty
        return ""
    
    return "\n".join(sheet_content)


def _extract_sheet_text_pandas(df, sheet_name: str) -> str:
    """Extract text from a DataFrame using pandas"""
    if df.empty:
        return ""
    
    sheet_content = []
    
    # Add sheet name as header
    if sheet_name and sheet_name.lower() not in ['sheet1', 'sheet', 'data']:
        sheet_content.append(f"=== {sheet_name} ===")
    
    # Analyze the DataFrame to identify text columns
    text_columns = _identify_text_columns_pandas(df)
    
    if not text_columns:
        # If no clear text columns, include all non-numeric content
        text_columns = _get_non_numeric_columns_pandas(df)
    
    # Extract text content
    for column in text_columns:
        column_values = df[column].dropna().astype(str)
        column_values = column_values[column_values.str.strip() != '']
        
        if len(column_values) > 0:
            # Add column header if meaningful
            if str(column) and str(column) != '0' and len(str(column)) > 1:
                sheet_content.append(f"--- {column} ---")
            
            # Add text values
            for value in column_values:
                cleaned_value = str(value).strip()
                if len(cleaned_value) >= 3 and _is_meaningful_text(cleaned_value):
                    sheet_content.append(cleaned_value)
    
    if len(sheet_content) <= 1:  # Only header or empty
        return ""
    
    return "\n".join(sheet_content)


def _analyze_sheet_content_openpyxl(sheet) -> Dict[str, Any]:
    """Analyze sheet content to understand structure"""
    analysis = {
        'has_headers': False,
        'text_columns': [],
        'total_rows': sheet.max_row or 0,
        'total_cols': sheet.max_column or 0
    }
    
    if analysis['total_rows'] < 2:
        return analysis
    
    # Check first row for headers
    first_row = [cell.value for cell in sheet[1]]
    second_row = [cell.value for cell in sheet[2]] if analysis['total_rows'] >= 2 else []
    
    # Detect headers based on content patterns
    if first_row and second_row:
        analysis['has_headers'] = _detect_headers_excel(first_row, second_row)
    
    # Identify text-heavy columns
    sample_size = min(10, analysis['total_rows'])
    for col_idx in range(1, analysis['total_cols'] + 1):
        col_values = []
        start_row = 2 if analysis['has_headers'] else 1
        
        for row_idx in range(start_row, min(start_row + sample_size, analysis['total_rows'] + 1)):
            cell_value = sheet.cell(row=row_idx, column=col_idx).value
            if cell_value is not None:
                col_values.append(str(cell_value))
        
        if _is_text_column_excel(col_values):
            analysis['text_columns'].append(col_idx)
    
    return analysis


def _extract_with_headers_openpyxl(sheet, analysis: Dict[str, Any]) -> List[str]:
    """Extract content from sheet with headers"""
    content = []
    
    # Get header row
    headers = [cell.value for cell in sheet[1]]
    
    # Extract text from identified text columns
    for col_idx in analysis['text_columns']:
        if col_idx <= len(headers):
            header_name = headers[col_idx - 1]
            
            if header_name and str(header_name).strip():
                content.append(f"--- {header_name} ---")
            
            # Extract column values
            for row_idx in range(2, analysis['total_rows'] + 1):
                cell_value = sheet.cell(row=row_idx, column=col_idx).value
                if cell_value is not None:
                    text_value = str(cell_value).strip()
                    if len(text_value) >= 3 and _is_meaningful_text(text_value):
                        content.append(text_value)
    
    return content


def _extract_without_headers_openpyxl(sheet, analysis: Dict[str, Any]) -> List[str]:
    """Extract content from sheet without headers"""
    content = []
    
    # Extract all text content from text columns
    for col_idx in analysis['text_columns']:
        col_content = []
        
        for row_idx in range(1, analysis['total_rows'] + 1):
            cell_value = sheet.cell(row=row_idx, column=col_idx).value
            if cell_value is not None:
                text_value = str(cell_value).strip()
                if len(text_value) >= 3 and _is_meaningful_text(text_value):
                    col_content.append(text_value)
        
        if col_content:
            content.extend(col_content)
    
    return content


def _extract_comments_openpyxl(sheet) -> List[str]:
    """Extract cell comments from the sheet"""
    comments = []
    
    try:
        for row in sheet.iter_rows():
            for cell in row:
                if cell.comment and cell.comment.text:
                    comment_text = cell.comment.text.strip()
                    if len(comment_text) >= 3:
                        comments.append(f"Comment: {comment_text}")
    except Exception:
        # Ignore comment extraction errors
        pass
    
    return comments


def _identify_text_columns_pandas(df) -> List[str]:
    """Identify columns containing meaningful text in a DataFrame"""
    text_columns = []
    
    for column in df.columns:
        sample_values = df[column].dropna().astype(str)
        sample_values = sample_values[sample_values.str.strip() != '']
        
        if len(sample_values) == 0:
            continue
        
        # Take a sample for analysis
        sample = sample_values.head(20).tolist()
        
        if _is_text_column_excel(sample):
            text_columns.append(column)
    
    return text_columns


def _get_non_numeric_columns_pandas(df) -> List[str]:
    """Get columns that are not purely numeric"""
    non_numeric_cols = []
    
    for column in df.columns:
        # Check if column contains meaningful non-numeric text
        sample_values = df[column].dropna().astype(str)
        non_numeric_count = 0
        
        for value in sample_values.head(20):
            if not _is_numeric_value(value) and len(str(value).strip()) >= 3:
                non_numeric_count += 1
        
        if non_numeric_count >= len(sample_values) * 0.3:  # At least 30% non-numeric
            non_numeric_cols.append(column)
    
    return non_numeric_cols


def _detect_headers_excel(first_row: List, second_row: List) -> bool:
    """Detect if first row contains headers"""
    if not first_row or not second_row:
        return True
    
    # Count text vs numeric in each row
    first_row_text = sum(1 for val in first_row if val and not _is_numeric_value(str(val)))
    second_row_text = sum(1 for val in second_row if val and not _is_numeric_value(str(val)))
    
    total_cols = len(first_row)
    
    if total_cols == 0:
        return True
    
    # If first row is mostly text and second row is mostly non-text, likely headers
    first_text_ratio = first_row_text / total_cols
    second_text_ratio = second_row_text / total_cols
    
    return first_text_ratio > 0.6 and second_text_ratio < 0.4


def _is_text_column_excel(sample_values: List[str]) -> bool:
    """Determine if a column contains meaningful text"""
    if not sample_values:
        return False
    
    text_count = 0
    
    for value in sample_values:
        value_str = str(value).strip()
        
        # Skip very short values
        if len(value_str) < 3:
            continue
        
        # Skip numeric values
        if _is_numeric_value(value_str):
            continue
        
        # Check for text characteristics
        has_letters = any(c.isalpha() for c in value_str)
        has_spaces = ' ' in value_str
        reasonable_length = 3 <= len(value_str) <= 1000
        
        if has_letters and reasonable_length:
            text_count += 1
    
    # Column is text if at least 40% of values are meaningful text
    return (text_count / len(sample_values)) >= 0.4 if sample_values else False


def _is_numeric_value(value: str) -> bool:
    """Check if a value is numeric"""
    value = str(value).strip()
    
    # Remove common formatting
    cleaned = value.replace(',', '').replace(' ', '').replace('%', '')
    cleaned = cleaned.replace('(', '-').replace(')', '')  # Negative numbers in parentheses
    
    try:
        float(cleaned)
        return True
    except ValueError:
        return False


def _is_meaningful_text(text: str) -> bool:
    """Check if text is meaningful (not just IDs, codes, etc.)"""
    text = text.strip()
    
    # Skip very short text
    if len(text) < 3:
        return False
    
    # Skip if it's purely numeric
    if _is_numeric_value(text):
        return False
    
    # Skip if it looks like an ID or code (all caps with numbers/underscores)
    if len(text) <= 10 and text.isupper() and any(c.isdigit() or c == '_' for c in text):
        return False
    
    # Skip if it's a simple date
    import re
    date_patterns = [
    r'^\d{4}-\d{2}-\d{2}$',
    r'^\d{2}/\d{2}/\d{4}$',
    r'^\d{2}-\d{2}-\d{4}$'
]

    if any(re.match(pattern, text) for pattern in date_patterns):
        return False
    
    # Must have at least one letter
    if not any(c.isalpha() for c in text):
        return False
    
    return True